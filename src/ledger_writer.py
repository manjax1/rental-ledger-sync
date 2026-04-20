from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_SHEET_NAME = "Ledger"
_DATA_START_ROW = 3

# Column indices (1-based): A–J
_COL_DATE     = 1
_COL_MONTH    = 2
_COL_MM_YYYY  = 3
_COL_YEAR     = 4
_COL_PROPERTY = 5
_COL_CATEGORY = 6
_COL_TYPE     = 7
_COL_DESC     = 8
_COL_AMOUNT   = 9
_COL_NOTES    = 10

_AMOUNT_FORMAT = r'\$#,##0.00_);[Red]"($"#,##0.00\);\-'

# (fill_hex, font_hex) per property
_PROPERTY_COLORS: dict[str, tuple[str, str]] = {
    "807 W. Center St Manteca CA 95337":          ("D6E4F7", "1F3864"),
    "809 W. Center St Manteca CA 95337":          ("DDEEFF", "1A4A78"),
    "2516 Mission Rd Stockton CA 95204":          ("D9EAD3", "1E3B14"),
    "2528 Mission Rd Stockton CA 95204":          ("E8F5D8", "2E4A1A"),
    "1867 Elmwood Ave Stockton CA 95204":         ("FCE8D5", "7B3F00"),
    "1867 Unit #A Elmwood Ave Stockton CA 95204": ("FDE9D9", "843D0B"),
    "16150 Sheltered Cove Lathrop CA 95330":      ("EAD7F5", "4B1E6B"),
    "16945 Strauss Court Lathrop CA 95330":       ("F3E6FF", "5C2D8A"),
    "2431 Mulholland Drive Lathrop CA 95330":     ("FADADD", "8B0000"),
    "1769 Mulholland Drive Lathrop CA 95330":     ("FDECEA", "7A1A1A"),
    "2 Cove Rd. Ponte Vedra Beach FL 32082":      ("D4F0EF", "004040"),
    "18 Meadow Star Ct Spring TX 77381":          ("D9EEF3", "0D4A56"),
    "6812 Dalmatian Circle Plano TX 75023":       ("FFF2CC", "4D3900"),
}
_DEFAULT_COLORS = ("F2F2F2", "444444")

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALIGN  = Alignment(horizontal="left", vertical="center")


def _to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _load_existing(ws) -> set[tuple]:
    """Build a dedup set of (date, round(amount,2), description) from existing rows."""
    keys: set[tuple] = set()
    for row in range(_DATA_START_ROW, ws.max_row + 1):
        date_val = ws.cell(row=row, column=_COL_DATE).value
        if date_val is None:
            continue
        amt_val  = ws.cell(row=row, column=_COL_AMOUNT).value
        desc_val = ws.cell(row=row, column=_COL_DESC).value
        try:
            d = _to_date(date_val)
        except (ValueError, TypeError):
            continue
        keys.add((d, round(float(amt_val or 0), 2), str(desc_val)))
    return keys


def _next_empty_row(ws) -> int:
    """Return the first row >= DATA_START_ROW whose Col A cell is empty."""
    for row in range(_DATA_START_ROW, ws.max_row + 2):
        if ws.cell(row=row, column=_COL_DATE).value is None:
            return row
    return ws.max_row + 2


def _apply_style(ws, row: int, fill_hex: str, font_hex: str):
    fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    font = Font(name="Arial", size=10, bold=False, color=font_hex)
    for col in range(1, _COL_NOTES + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = _ALIGN
        cell.border    = _BORDER


def write_transactions(matched_transactions: list[dict], ledger_file_path: str) -> dict:
    """
    Append matched transactions to the Ledger sheet of the Excel file.

    Skips any transaction where (date, amount, description) already exists.
    Returns {"added": int, "skipped": int, "total": int}.
    """
    path = Path(ledger_file_path)
    wb = load_workbook(path)
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{_SHEET_NAME}' not found in {path}")
    ws = wb[_SHEET_NAME]

    existing          = _load_existing(ws)
    added             = 0
    skipped           = 0
    added_transactions: list[dict] = []

    for txn in matched_transactions:
        txn_date    = _to_date(txn["date"])
        raw_amount  = float(txn["amount"])
        # Income: always store as positive. Expense: BofA exports are already positive.
        amount      = round(abs(raw_amount) if txn.get("type") == "Income" else abs(raw_amount), 2)
        description = txn["name"]
        dedup_key   = (txn_date, amount, description)

        if dedup_key in existing:
            skipped += 1
            continue

        row          = _next_empty_row(ws)
        property_val = txn.get("property", "")
        fill_hex, font_hex = _PROPERTY_COLORS.get(property_val, _DEFAULT_COLORS)

        # --- Col A: Date ---
        c = ws.cell(row=row, column=_COL_DATE)
        c.value         = txn_date
        c.number_format = "MM-DD-YYYY"

        # --- Col B: Month # ---
        c = ws.cell(row=row, column=_COL_MONTH)
        c.value         = txn_date.month
        c.number_format = "0"

        # --- Col C: MM-YYYY ---
        ws.cell(row=row, column=_COL_MM_YYYY).value = txn_date.strftime("%m-%Y")

        # --- Col D: Year ---
        c = ws.cell(row=row, column=_COL_YEAR)
        c.value         = txn_date.year
        c.number_format = "0"

        # --- Col E–H: Property / Category / Type / Description ---
        ws.cell(row=row, column=_COL_PROPERTY).value = property_val
        ws.cell(row=row, column=_COL_CATEGORY).value = txn.get("category", "")
        ws.cell(row=row, column=_COL_TYPE).value     = txn.get("type", "")
        ws.cell(row=row, column=_COL_DESC).value     = description

        # --- Col I: Amount ---
        c = ws.cell(row=row, column=_COL_AMOUNT)
        c.value         = amount
        c.number_format = _AMOUNT_FORMAT

        # --- Col J: Notes ---
        ws.cell(row=row, column=_COL_NOTES).value = txn.get("note", "")

        # Apply fill, font, alignment, border to all 10 columns
        _apply_style(ws, row, fill_hex, font_hex)

        existing.add(dedup_key)
        added_transactions.append(txn)
        added += 1

    if added:
        wb.save(path)

    return {
        "added":             added,
        "skipped":           skipped,
        "total":             added + skipped,
        "added_transactions": added_transactions,
    }


def update_mortgage_category_in_ledger(ledger_path: str) -> None:
    """
    One-time migration: rename "Mortgage - Principal" → "Mortgage - P + I"
    in the hidden Categories sheet and all Ledger rows (col F).
    """
    OLD = "Mortgage - Principal"
    NEW = "Mortgage - P + I"

    path = Path(ledger_path)
    wb   = load_workbook(path)

    # --- Categories sheet ---
    categories_updated = False
    if "Categories" in wb.sheetnames:
        ws_cat = wb["Categories"]
        for row in ws_cat.iter_rows():
            for cell in row:
                if cell.value == OLD:
                    cell.value = NEW
                    categories_updated = True

    # --- Ledger sheet ---
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{_SHEET_NAME}' not found in {path}")
    ws      = wb[_SHEET_NAME]
    updated = 0
    for row in range(_DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=row, column=_COL_CATEGORY)
        if cell.value == OLD:
            cell.value = NEW
            updated += 1

    wb.save(path)

    if categories_updated:
        print(f"✅ Updated Categories sheet: {OLD} → {NEW}")
    else:
        print(f"⚠️  '{OLD}' not found in Categories sheet (may already be updated)")
    print(f"✅ Updated {updated} ledger rows with new mortgage category")


def fix_transaction_category(
    ledger_path: str,
    date_str: str,
    amount: float,
    old_category: str,
    new_category: str,
) -> None:
    """
    One-time fix: update Col F (Category) for a specific transaction row.

    Matches on:
      Col A — date equals date_str (YYYY-MM-DD)
      Col I — amount equals amount (rounded to 2 decimal places)
      Col F — category equals old_category

    Raises ValueError if no matching row is found.
    """
    path        = Path(ledger_path)
    wb          = load_workbook(path)
    if _SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{_SHEET_NAME}' not found in {path}")
    ws          = wb[_SHEET_NAME]
    target_date = _to_date(date_str)
    target_amt  = round(float(amount), 2)

    for row in range(_DATA_START_ROW, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=_COL_DATE).value
        if cell_date is None:
            continue
        try:
            row_date = _to_date(cell_date)
        except (ValueError, TypeError):
            continue

        row_amt      = round(float(ws.cell(row=row, column=_COL_AMOUNT).value or 0), 2)
        row_category = ws.cell(row=row, column=_COL_CATEGORY).value

        if row_date == target_date and row_amt == target_amt and row_category == old_category:
            ws.cell(row=row, column=_COL_CATEGORY).value = new_category
            wb.save(path)
            print(f"✅ Fixed category for transaction on {date_str}: {old_category} → {new_category}")
            return

    raise ValueError(
        f"No matching row found: date={date_str}, amount={amount}, category='{old_category}'"
    )
